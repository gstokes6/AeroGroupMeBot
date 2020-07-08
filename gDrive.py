from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
import os
import datetime
import openpyxl

import LOAD_ENV_VARS

class gDrive:
    def __init__(self):
        ##vars to use later
        self.lastScheduleUpdateTime = datetime.datetime.min
        self.scheduleData = []
        
        ##load important stuff
        self.root = LOAD_ENV_VARS.ENV_VARS['root']
        self.access_token = LOAD_ENV_VARS.ENV_VARS['gd_access_token']
        self.client_secret = LOAD_ENV_VARS.ENV_VARS['gd_client_secret']
        self.client_id = LOAD_ENV_VARS.ENV_VARS['gd_client_id']
        self.refresh_token = LOAD_ENV_VARS.ENV_VARS['gd_refresh_token']
        self.token_expiry = LOAD_ENV_VARS.ENV_VARS['gd_token_expiry']
        if True:
            ##make client creds
            Text = """{"access_token": %s, "client_id": %s, "client_secret": %s, "refresh_token": %s, "token_expiry": %s, "token_uri": "https://oauth2.googleapis.com/token", "user_agent": null, "revoke_uri": "https://oauth2.googleapis.com/revoke", "id_token": null, "id_token_jwt": null, "token_response": {"access_token": %s, "expires_in": 3600, "refresh_token": %s, "scope": "https://www.googleapis.com/auth/drive", "token_type": "Bearer"}, "scopes": ["https://www.googleapis.com/auth/drive"], "token_info_uri": "https://oauth2.googleapis.com/tokeninfo", "invalid": false, "_class": "OAuth2Credentials", "_module": "oauth2client.client"}"""%(self.access_token,self.client_id,self.client_secret,self.refresh_token,self.token_expiry,self.access_token,self.refresh_token)
            f = open("mycreds.txt","w+")
            f.write(Text)
            f.close()
            ##make client secrets from enviromental variables
            Text = '''{"installed":{"client_id":%s,"project_id":"quickstart-1564436220867","auth_uri":"https://accounts.google.com/o/oauth2/auth","token_uri":"https://oauth2.googleapis.com/token","auth_provider_x509_cert_url":"https://www.googleapis.com/oauth2/v1/certs","client_secret":%s,"redirect_uris":["urn:ietf:wg:oauth:2.0:oob","http://localhost"]}}'''%(self.client_id,self.client_secret)
            f = open("client_secrets.json","w+")
            f.write(Text)
            f.close()
        else:
            self.root = 'Python Bot Test'
        
        # Try to load saved client credentials
        gauth = GoogleAuth()
        gauth.LoadCredentialsFile("mycreds.txt")
        if gauth.credentials is None:
            # Authenticate if they're not there
            gauth.LocalWebserverAuth()
        elif gauth.access_token_expired:
            # Refresh them if expired
            gauth.Refresh()
        else:
            # Initialize the saved creds
            gauth.Authorize()
        # Save the current credentials to a file
        gauth.SaveCredentialsFile("mycreds.txt")
        self.UpdateEnvVars()
        #initialize drive object
        self.drive = GoogleDrive(gauth)
        ##Setup sechdule
        self.loadSchedule()

    def loadSchedule(self):
        self.lastScheduleUpdateTime = datetime.datetime.now()
        SpreadsheetFile = self.FindOrCreateFolder( ['Classes.xlsx'] )
        
        Spreadsheet = self.drive.CreateFile( { 'id':SpreadsheetFile['id'] } )
        Spreadsheet.GetContentFile('Classes.xlsx')
        wb = openpyxl.load_workbook('Classes.xlsx')
        DataSheet = wb['Classes']
        Row = 2
        self.scheduleData = []
        while DataSheet.cell(row = Row, column = 1).value:
            classDays = DataSheet.cell(row = Row, column = 2).value
            struct = {
                'className':(DataSheet.cell(row = Row, column = 1).value),
                'startTime':(DataSheet.cell(row = Row, column = 3).value),
                'endTime':(DataSheet.cell(row = Row, column = 4).value),
                'classDays':classDays.replace('Su','0').replace('M','1').replace('Tu','2').replace('W','3').replace('Th','4').replace('F','5').replace('Sa','6')
                }
            self.scheduleData.append(struct)
            Row = Row + 1

    def checkClasses(self,message):
        now = datetime.datetime.fromtimestamp(message['created_at'])
        timeNow = now.time()
        dateNow = now.date()
        print(self.scheduleData)
        print(now)
        if ( (now-self.lastScheduleUpdateTime) > datetime.timedelta(days=0,hours=1,minutes=0) ):
            self.loadSchedule()

        messageScheduleList = []
        for Class in self.scheduleData:
            inClassTime = ( (Class['startTime']<timeNow) and (timeNow<Class['endTime']) )
            onClassDay = ( str(dateNow.weekday()) in Class['classDays'] )
            print( str( dateNow.weekday( ) ) )
            print(  str(dateNow.weekday()) == classDays)
            print(inClassTime,onClassDay)
            if inClassTime and onClassDay:
                messageScheduleList.append(Class['className'])
            
        return messageScheduleList
        
    def UpdateEnvVars(self):
        f = open('mycreds.txt','r')
        c = f.read()
        s = c.split('"')
        self.gd_access_token = s[3]
        self.gd_client_secret = s[11]
        self.gd_client_id = s[7]
        self.gd_refresh_token = s[15]
        self.gd_token_expiry = s[19]
        os.environ["GD_ACCESS_TOKEN"] = self.gd_access_token
        os.environ["GD_CLIENT_SECRET"] = self.gd_client_secret
        os.environ["GD_CLIENT_ID"] = self.gd_client_id
        os.environ["GD_REFRESH_TOKEN"] = self.gd_refresh_token
        os.environ["GD_TOKEN_EXPIRY"] = self.gd_token_expiry

    def FindOrCreateFolder(self,folderNames):
        parent_id = 'root'
        folderNames.insert(0,self.root)
        for folderName in folderNames:
            search_list = []
            file_list = self.drive.ListFile({'q': "'%s' in parents and trashed=false"%(parent_id)}).GetList()
            for file in file_list:
                    if (file['title'] == folderName):
                            search_list.append(file)
            if len(search_list) == 0:
                # Create folder.
                folder_metadata = {
                    'title' : folderName,
                    # The mimetype defines this new file as a folder, so don't change this.
                    'mimeType' : 'application/vnd.google-apps.folder',
                    "parents"  : [{"kind": "drive#fileLink", "id": parent_id}]
                }
                folder = self.drive.CreateFile(folder_metadata)
                folder.Upload()
                
            else:
                folder = search_list[0]
            ##now look in that folder
            parent_id = folder['id']
        return (folder)
    
    def UploadFile(self,path,folderNames):
        folder = self.FindOrCreateFolder(fodlerNames)
        f = self.drive.CreateFile({"parents": [{"kind": "drive#fileLink", "id": folder['id']}],'title': path})
        f.SetContentFile(path)
        f.Upload()
        print('Uploaded!')
        

