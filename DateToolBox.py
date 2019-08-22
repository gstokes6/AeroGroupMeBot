import time
import datetime
import openpyxl
import GoogleDrive as GD
import re

def GetSpreadsheet(Path):
    wb = openpyxl.load_workbook(Path)
    Datasheet = wb['Classes']
    return Datasheet,wb

def GetSpreadsheetContents(DataSheet):
    Classes = []
    MeetingDays = []
    TimeStart = []
    TimeEnd = []
    Row = 2
    while DataSheet.cell(row = Row, column = 1).value:
        Classes.append(DataSheet.cell(row = Row, column = 1).value)
        MeetingDays.append(DataSheet.cell(row = Row, column = 2).value)
        TimeStart.append(DataSheet.cell(row = Row, column = 3).value)
        TimeEnd.append(DataSheet.cell(row = Row, column = 4).value)
        Row = Row + 1
    return Classes,MeetingDays,TimeStart,TimeEnd

def UpdateDataSheet(drive,root):
    SpreadsheetID = GD.FindOrCreateFolder(drive,[root,'Classes.xlsx'])
    Spreadsheet = drive.CreateFile({'id':SpreadsheetID})
    Spreadsheet.GetContentFile('Classes.xlsx')
    DataSheet,wb = GetSpreadsheet('Classes.xlsx')
    return DataSheet
    
def AddClassFolders(drive,root):
    DataSheet = UpdateDataSheet(drive,root)
    Classes,MeetingDays,TimeStart,TimeEnd = GetSpreadsheetContents(DataSheet)
    for Class in Classes:
        GD.FindOrCreateFolder(drive,['Python Bot','Uploads',Class])

def IsInClass(drive,root,timestamp):
    DataSheet = UpdateDataSheet(drive)
    Classes,MeetingDays,TimeStart,TimeEnd = GetSpreadsheetContents(DataSheet)
    Time = datetime.datetime.fromtimestamp(timestamp)
    AdjMeeting = []
    AdjStart = []
    AdjEnd = []
    for i in range(0,len(Classes)):
        temp = re.findall('[A-Z][a-z]*', MeetingDays[i])
        for day in temp:
            if day == 'M':
                temp[temp.index(day)] = 0
            elif (day=='T') or (day=='Tu'):
                temp[temp.index(day)] = 1
            elif (day=='W'):
                temp[temp.index(day)] = 2
            elif (day=='Th'):
                temp[temp.index(day)] = 3
            elif (day=='F'):
                temp[temp.index(day)] = 4
            elif (day=='Sa'):
                temp[temp.index(day)] = 5
            elif (day=='Su'):
                temp[temp.index(day)] = 6
            else:
                temp[temp.index(day)] = 7
        AdjMeeting.append(temp)
    for i in range(0,len(Classes)):
        if (Time.weekday() in AdjMeeting[i]) and (TimeStart[i] < Time.time() < TimeEnd[i]):
            return Classes[i]
    return None
    


if __name__ == "__main__":
    drive = GD.GetDrive()
    DataSheet = UpdateDataSheet(drive)
    print(DataSheet)























    
